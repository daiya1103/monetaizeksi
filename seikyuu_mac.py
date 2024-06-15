import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os
import pathlib
import sys

import tkinter as tk
from tkinter import BOTTOM, LEFT, ttk, StringVar, BooleanVar, IntVar, filedialog
from datetime import date

CUR_DIR = str(pathlib.Path(sys.argv[0]).resolve().parent)
NAME = '請求書の宛名'
BILLING = '請求合計'
SHEET_NAME = '触る用'

w = 400
h = 300

class Main(tk.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=w, height=h)
        self.root = root
        self.pack()
        self.widgets = []
        self.main_widgets()

    def main_widgets(self):
        self.select_frame = tk.Frame(self)
        self.select_frame.pack(pady=35)
        self.category_v = StringVar()
        self.category_v.set("出展者用")
        self.category_select = ttk.Combobox(
            self.select_frame,
            values=['出展者用', '運営用'],
            textvariable=self.category_v,
            state="readonly",
        )
        self.category_select.pack()
        self.excel_frame = tk.Frame(self)
        self.excel_frame.pack(pady=(0, 20))
        self.excel_label = tk.Label(self.excel_frame, text='請求先エクセルデータ')
        self.excel_label.pack()
        self.excel_v = StringVar()
        self.excel_input = ttk.Entry(self.excel_frame, textvariable=self.excel_v)
        self.excel_input.pack(side=LEFT)
        self.excel_button = ttk.Button(
            self.excel_frame, text="参照", command=lambda: self.filedialog_clicked()
        )
        self.excel_button.pack(side=LEFT)
        self.read_button = ttk.Button(
            self.excel_frame, text="読み込み", command=lambda: self.read_excel()
        )
        self.read_button.pack(side=LEFT)

        self.content_v = tk.StringVar()
        self.content_v.set([])

        self.content_frame = tk.Frame(self)
        self.content_frame.pack(pady=(0,5))
        self.content_label = tk.Label(self.content_frame, text='内訳対象')
        self.content_label.pack()
        self.content_box = tk.Listbox(self.content_frame, listvariable=self.content_v, height=5, selectmode=tk.MULTIPLE)
        self.content_box.pack()

        self.create_frame = tk.Frame(self)
        self.create_frame.pack()
        self.create_btn = ttk.Button(self.create_frame, text='請求書作成開始', command=lambda: self.create())
        self.create_btn.pack(side=LEFT)

    def filedialog_clicked(self):
        fTyp = [("", "*")]
        iFile = CUR_DIR + "/original"
        iFilePath = filedialog.askopenfilename(filetypes=fTyp, initialdir=iFile)
        self.excel_v.set(iFilePath)

    def read_excel(self):
        TARGET = self.excel_v.get()
        df = pd.read_excel(TARGET, header=1, sheet_name=SHEET_NAME)
        cand = df.columns.tolist()
        self.content_v.set(cand)

    def create(self):
        dt = date.today()
        date_all = dt.strftime('%Y/%m/%d')
        date_str = str(dt.year)[2:4] + str(dt.month) + str(dt.day)
        TARGET = self.excel_v.get()
        MODE = self.category_v.get()
        df = pd.read_excel(TARGET, header=1, sheet_name=SHEET_NAME)
        contents = self.content_box.curselection()
        seleted_items = []
        if not contents:
            print('not selected')
            return
        else:
            """
            複数のアイテムが選択される場合があるので、選択されたリストボックスのアイテムのインデックスをタプルのループを回し、インデックス毎にアイテムを取得する。
            ここでは、取得しアイテムをリストに保存するようにしておいたけど、勿論、リストに保存しなくてもいい。
            選択されたリストボックスのアイテムをリスト化するのは、なんか、他にもっといい書き方がありそうだけど…(´･ω･`)
            ま、いっか。
            """
            seleted_items = []
            for index in contents:
                listbox_item = self.content_box.get(index)
                seleted_items.append(listbox_item)
        
        df = pd.read_excel(TARGET, header=1, sheet_name=SHEET_NAME)
        sakkas = df[[NAME, BILLING] + seleted_items].to_dict(orient='records')

        CONTENT_ROW = 0
        NAME_CELL = 'A0'
        BILLING_CELL = 'A0'
        DATE_CELL = 'A0'
        XLSX_DIR = CUR_DIR + '/data/date_name_billing_出展者宛いきものづくし大規模請求書.xlsx'


        if MODE == '出展者用':
            XLSX_DIR = CUR_DIR + '/data/date_name_billing_出展者宛いきものづくし大規模請求書.xlsx'
            NAME_CELL = 'B6'
            BILLING_CELL = 'B20'
            DATE_CELL = 'G4'
            CONTENT_ROW = 24
        elif MODE == '運営用':
            XLSX_DIR = CUR_DIR + '/data/date_name_billing_運営宛いきものづくしガチャ請求書.xlsx'
            NAME_CELL = 'B8'
            BILLING_CELL = 'B17'
            DATE_CELL = 'G6'
            CONTENT_ROW = 23
        else: raise Exception

        for sakka in sakkas:
            try: 
                wb = load_workbook(XLSX_DIR)
                ws = wb.active
                if MODE == '出展者用':
                    ws[NAME_CELL] = sakka[NAME]
                ws[BILLING_CELL] = int(sakka[BILLING])
                ws[DATE_CELL] = date_all

                for i, content in enumerate(seleted_items):
                    ws.insert_rows(CONTENT_ROW+i)
                    ws.merge_cells(f'B{CONTENT_ROW+i}:C{CONTENT_ROW+i}')
                    ws[f'A{CONTENT_ROW+i}'] = content
                    ws[f'B{CONTENT_ROW+i}'] = int(sakka[content])
                    ws[f'D{CONTENT_ROW+i}'] = '円'
                    ws[f'A{CONTENT_ROW+i}'].border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws[f'B{CONTENT_ROW+i}'].border = Border(
                        left=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws[f'C{CONTENT_ROW+i}'].border = Border(
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws[f'D{CONTENT_ROW+i}'].border = Border(
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                if MODE == '出展者用':
                    OUTPUT_DIR = f'{CUR_DIR}/output/{date_str}_{sakka[NAME]}_{int(sakka[BILLING])}円_出展者宛いきものづくし大規模請求書.xlsx'
                    wb.save(OUTPUT_DIR)
                elif MODE == '運営用':
                    OUTPUT_DIR = f'{CUR_DIR}/output/{date_str}_{sakka[NAME]}_{int(sakka[BILLING])}円_運営宛いきものづくしガチャ請求書.xlsx'
                    wb.save(OUTPUT_DIR)

            except: print(sakka)


root = tk.Tk()
root.title('Monetize(ξ)')
root_geo = f"{w}x{h}"
root.geometry(root_geo)
root.iconbitmap(CUR_DIR + '/icon.ico')
app = Main(root=root)
app.mainloop()