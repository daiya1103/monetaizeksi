import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os
import pathlib
import sys

import tkinter as tk
import tkinter.simpledialog as simpledialog
from tkinter import BOTTOM, LEFT, ttk, StringVar, BooleanVar, IntVar, filedialog, messagebox
from datetime import date
import datetime

import glob
import xlwings as xw

import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import json

CUR_DIR = str(pathlib.Path(sys.argv[0]).resolve().parent)
NAME = '請求書の宛名'
BILLING = '請求合計'
EMAIL_COLUMN = 'メールアドレス'
SHEET_NAME = '触る用'

w = 400
h = 320

class Main(tk.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=w, height=h)
        self.root = root
        self.pack()
        self.widgets = []
        self.main_widgets()

    def main_widgets(self):
        self.select_frame = tk.Frame(self)
        self.select_frame.pack(pady=20)
        self.category_v = StringVar()
        self.category_v.set("出展者用")
        self.category_select = ttk.Combobox(
            self.select_frame,
            values=['出展者用', '運営用'],
            textvariable=self.category_v,
            state="readonly",
        )
        self.category_select.pack()
        self.excel_frame = tk.Frame(self)
        self.excel_frame.pack(pady=(0, 20))
        self.excel_label = tk.Label(self.excel_frame, text='請求先エクセルデータ')
        self.excel_label.pack()
        self.excel_v = StringVar()
        self.excel_input = ttk.Entry(self.excel_frame, textvariable=self.excel_v)
        self.excel_input.pack(side=LEFT)
        self.excel_button = ttk.Button(
            self.excel_frame, text="参照", command=lambda: self.filedialog_clicked()
        )
        self.excel_button.pack(side=LEFT)
        self.read_button = ttk.Button(
            self.excel_frame, text="読み込み", command=lambda: self.read_excel()
        )
        self.read_button.pack(side=LEFT)

        self.content_v = tk.StringVar()
        self.content_v.set([])

        self.content_frame = tk.Frame(self)
        self.content_frame.pack(pady=(0,5))
        self.content_label = tk.Label(self.content_frame, text='内訳対象')
        self.content_label.pack()
        self.content_box = tk.Listbox(self.content_frame, listvariable=self.content_v, height=5, selectmode=tk.MULTIPLE)
        self.content_box.pack()

        self.create_frame = tk.Frame(self)
        self.create_frame.pack()
        self.sample_btn = ttk.Button(self.create_frame, text='サンプル出力', command=lambda: self.sample())
        self.sample_btn.pack()
        self.create_btn = ttk.Button(self.create_frame, text='請求書作成開始', command=lambda: self.create())
        self.create_btn.pack()
        self.mail_btn = ttk.Button(self.create_frame, text='メール送信', command=lambda: self.send_email())
        self.mail_btn.pack()

    def filedialog_clicked(self):
        fTyp = [("", "*")]
        iFile = CUR_DIR + "/original"
        iFilePath = filedialog.askopenfilename(filetypes=fTyp, initialdir=iFile)
        self.excel_v.set(iFilePath)

    def read_excel(self):
        TARGET = self.excel_v.get()
        df = pd.read_excel(TARGET, header=1, sheet_name=SHEET_NAME)
        cand = df.columns.tolist()
        self.content_v.set(cand)

    def create(self):
        seleted_items = []
        contents = self.content_box.curselection()
        if not contents:
            res = messagebox.showerror("てへぺろ", "内訳の選択がないです。")
            return
        else:
            """
            複数のアイテムが選択される場合があるので、選択されたリストボックスのアイテムのインデックスをタプルのループを回し、インデックス毎にアイテムを取得する。
            ここでは、取得しアイテムをリストに保存するようにしておいたけど、勿論、リストに保存しなくてもいい。
            選択されたリストボックスのアイテムをリスト化するのは、なんか、他にもっといい書き方がありそうだけど…(´･ω･`)
            ま、いっか。
            """
            seleted_items = []
            for index in contents:
                listbox_item = self.content_box.get(index)
                seleted_items.append(listbox_item)
        print('-----------------------------')
        print('請求書発行スタート')
        print(f'内訳: {seleted_items}')
        print('-----------------------------')
        dt = date.today()
        date_all = dt.strftime('%Y/%m/%d')
        month = str(dt.month)
        if len(month) == 1:
            month = '0' + month
        date_A = str(dt.day)
        if len(date_A) == 1:
            date_A = '0' + date_A
        date_str = str(dt.year)[2:4] + month + date_A
        TARGET = self.excel_v.get()
        MODE = self.category_v.get()
        df = pd.read_excel(TARGET, header=1, sheet_name=SHEET_NAME)
        sakkas = df[[NAME, BILLING, EMAIL_COLUMN] + seleted_items].to_dict(orient='records')

        CONTENT_ROW = 0
        NAME_CELL = 'A0'
        BILLING_CELL = 'A0'
        DATE_CELL = 'A0'
        XLSX_DIR = CUR_DIR + '/data/date_name_billing_出展者宛いきものづくし大規模請求書.xlsx'

        if MODE == '出展者用':
            XLSX_DIR = CUR_DIR + '/data/date_name_billing_出展者宛いきものづくし大規模請求書.xlsx'
            NAME_CELL = 'B6'
            BILLING_CELL = 'B20'
            DATE_CELL = 'G4'
            CONTENT_ROW = 24
        elif MODE == '運営用':
            XLSX_DIR = CUR_DIR + '/data/date_name_billing_運営宛いきものづくしガチャ請求書.xlsx'
            NAME_CELL = 'B8'
            BILLING_CELL = 'B17'
            DATE_CELL = 'G6'
            CONTENT_ROW = 23
        else: raise Exception

        App = xw.App(visible=False)

        for i, sakka in enumerate(sakkas):
            print(f"処理中…({i+1}/{len(sakkas)}人)", end="")
            sum = 0
            for content in seleted_items:
                sum += int(sakkas[0][content])
            if sum == int(sakkas[0][BILLING]):
                try: 
                    wb = load_workbook(XLSX_DIR)
                    ws = wb.active
                    if MODE == '出展者用':
                        ws[NAME_CELL] = sakka[NAME]
                    ws[BILLING_CELL] = int(sakka[BILLING])
                    ws[DATE_CELL] = date_all

                    for i, content in enumerate(seleted_items):
                        ws.insert_rows(CONTENT_ROW+i)
                        ws.merge_cells(f'B{CONTENT_ROW+i}:C{CONTENT_ROW+i}')
                        ws[f'A{CONTENT_ROW+i}'] = content
                        ws[f'B{CONTENT_ROW+i}'] = int(sakka[content])
                        ws[f'D{CONTENT_ROW+i}'] = '円'
                        ws[f'A{CONTENT_ROW+i}'].border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        ws[f'B{CONTENT_ROW+i}'].border = Border(
                            left=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        ws[f'C{CONTENT_ROW+i}'].border = Border(
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        ws[f'D{CONTENT_ROW+i}'].border = Border(
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    if MODE == '出展者用':
                        OUTPUT_DIR = f'{CUR_DIR}/output/{date_str}_{sakka[NAME]}_{int(sakka[BILLING])}円_出展者宛いきものづくし大規模請求書.xlsx'
                        wb.save(OUTPUT_DIR)
                        PDF_PATH = OUTPUT_DIR.replace('.xlsx', '.pdf')
                        xwb = xw.Book(OUTPUT_DIR)
                        xwb.to_pdf(path=PDF_PATH, include=None, exclude=None, exclude_start_string='#', show=False)
                        xwb.close()
                        os.remove(OUTPUT_DIR)
                        sakka['pdf'] = PDF_PATH
                        sakkas[i] = sakka
                    elif MODE == '運営用':
                        OUTPUT_DIR = f'{CUR_DIR}/output/{date_str}_{sakka[NAME]}_{int(sakka[BILLING])}円_運営宛いきものづくしガチャ請求書.xlsx'
                        wb.save(OUTPUT_DIR)
                    print('\r', end='')

                except Exception as e:
                    if sakka[NAME] != 'nan':
                        print(e)
                        print(f'{sakka[NAME]}の出力に失敗しました。')

            else:
                res = messagebox.showerror("てへぺろ", "内訳の合計と請求金額が違う。彩度選択してみてください")
                print('-----------------------------')
                print('処理を中止しました。')
                print('-----------------------------')
                return
            
        dt6= datetime.datetime.now()
        strdt6 = dt6.strftime('%Y-%m-%d %H:%M:%S')
        json_dict = {'sakkas': sakkas, 'time': strdt6}
        
        with open(f'{CUR_DIR}/mail.json', 'w') as f:
            json.dump(json_dict, f)
        print('-----------------------------')
        print('全セラーの処理が完了しました。')
        print('-----------------------------')
        print('\n')

        App.quit()

    def sample(self):
        seleted_items = []
        contents = self.content_box.curselection()
        if not contents:
            res = messagebox.showerror("てへぺろ", "内訳の選択がないです。")
            return
        else:
            """
            複数のアイテムが選択される場合があるので、選択されたリストボックスのアイテムのインデックスをタプルのループを回し、インデックス毎にアイテムを取得する。
            ここでは、取得しアイテムをリストに保存するようにしておいたけど、勿論、リストに保存しなくてもいい。
            選択されたリストボックスのアイテムをリスト化するのは、なんか、他にもっといい書き方がありそうだけど…(´･ω･`)
            ま、いっか。
            """
            seleted_items = []
            for index in contents:
                listbox_item = self.content_box.get(index)
                seleted_items.append(listbox_item)
        print('-----------------------------')
        print('サンプル出力')
        print(f'内訳: {seleted_items}')
        print('-----------------------------')
        dt = date.today()
        date_all = dt.strftime('%Y/%m/%d')
        month = str(dt.month)
        if len(month) == 1:
            month = '0' + month
        date_A = str(dt.day)
        if len(date_A) == 1:
            date_A = '0' + date_A
        date_str = str(dt.year)[2:4] + month + date_A
        TARGET = self.excel_v.get()
        MODE = self.category_v.get()
        df = pd.read_excel(TARGET, header=1, sheet_name=SHEET_NAME)
        sakkas = df[[NAME, BILLING] + seleted_items].to_dict(orient='records')

        CONTENT_ROW = 0
        NAME_CELL = 'A0'
        BILLING_CELL = 'A0'
        DATE_CELL = 'A0'
        XLSX_DIR = CUR_DIR + '/data/date_name_billing_出展者宛いきものづくし大規模請求書.xlsx'

        if MODE == '出展者用':
            XLSX_DIR = CUR_DIR + '/data/date_name_billing_出展者宛いきものづくし大規模請求書.xlsx'
            NAME_CELL = 'B6'
            BILLING_CELL = 'B20'
            DATE_CELL = 'G4'
            CONTENT_ROW = 24
        elif MODE == '運営用':
            XLSX_DIR = CUR_DIR + '/data/date_name_billing_運営宛いきものづくしガチャ請求書.xlsx'
            NAME_CELL = 'B8'
            BILLING_CELL = 'B17'
            DATE_CELL = 'G6'
            CONTENT_ROW = 23
        else: raise Exception

        App = xw.App()
        try:
            sum = 0
            for content in seleted_items:
                sum += int(sakkas[0][content])
            
            if sum == int(sakkas[0][BILLING]):
                wb = load_workbook(XLSX_DIR)
                ws = wb.active
                if MODE == '出展者用':
                    ws[NAME_CELL] = sakkas[0][NAME]
                ws[BILLING_CELL] = int(sakkas[0][BILLING])
                ws[DATE_CELL] = date_all

                for i, content in enumerate(seleted_items):
                    ws.insert_rows(CONTENT_ROW+i)
                    ws.merge_cells(f'B{CONTENT_ROW+i}:C{CONTENT_ROW+i}')
                    ws[f'A{CONTENT_ROW+i}'] = content
                    ws[f'B{CONTENT_ROW+i}'] = int(sakkas[0][content])
                    ws[f'D{CONTENT_ROW+i}'] = '円'
                    ws[f'A{CONTENT_ROW+i}'].border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws[f'B{CONTENT_ROW+i}'].border = Border(
                        left=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws[f'C{CONTENT_ROW+i}'].border = Border(
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws[f'D{CONTENT_ROW+i}'].border = Border(
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                if MODE == '出展者用':
                    OUTPUT_DIR = f'{CUR_DIR}/output/{date_str}_{sakkas[0][NAME]}_{int(sakkas[0][BILLING])}円_出展者宛いきものづくし大規模請求書.xlsx'
                    wb.save(OUTPUT_DIR)
                elif MODE == '運営用':
                    OUTPUT_DIR = f'{CUR_DIR}/output/{date_str}_{sakkas[0][NAME]}_{int(sakkas[0][BILLING])}円_運営宛いきものづくしガチャ請求書.xlsx'
                    wb.save(OUTPUT_DIR)
                xwb = xw.Book(OUTPUT_DIR)
                print('\r', end='')
            else:
                res = messagebox.showerror("てへぺろ", "内訳の合計と請求金額が違う。再度選択してみてください")
                return

        except Exception as e:
            print(e)
            if sakkas[0][NAME] != 'nan':
                print(f'{sakkas[0][NAME]}の出力に失敗しました。')

    def send_email(self):
        with open(f'{CUR_DIR}/mail.json') as f:
            d = json.load(f)
        print(d['time'])
        sakkas = d['sakkas']
        res = messagebox.askquestion('メールの送信', f'{str(d['time'])}に作成したリストにメールを送りますか？')
        if res == 'yes':
            GMAIL_PASSWORD = 'wrmzbegoyhjzvlfw'
            GMAIL_ADDRESS = 'tsukushi.ikimono.team@gmail.com'
            # 送信先のアドレスを登録します
            send_address = "doraemonyeah666@gmail.com"


            # 件名、送信先アドレス、本文を渡す関数です
            def make_mime_text(mail_to, subject):
                msg = MIMEMultipart()
                msg["Subject"] = subject
                msg["To"] = mail_to
                msg["From"] = GMAIL_ADDRESS
                return msg

            # smtp経由でメール送信する関数です
            def send_gmail(msg):
                server = smtplib.SMTP_SSL(
                    "smtp.gmail.com", 465,
                    context = ssl.create_default_context())
                server.set_debuglevel(0)
                server.login(GMAIL_ADDRESS, GMAIL_PASSWORD)
                server.send_message(msg)
            msg = make_mime_text(
                mail_to = send_address,
                subject = "テスト送信",
            )
            body_file_path = 'path_to_body.txt'
            with open(body_file_path, 'r', encoding='utf-8') as file:
                body = file.read()
            body = "Pythonでのメール送信です"
            msg.attach(MIMEText(body, 'plain'))
            pdf_path = CUR_DIR + '/output/240608_Hamoon_34100円_出展者宛いきものづくし大規模請求書.pdf'
            with open(pdf_path, 'rb') as pdf_file:
                mime_base = MIMEBase('application', 'octet-stream')
                mime_base.set_payload(pdf_file.read())
                encoders.encode_base64(mime_base)
                mime_base.add_header('Content-Disposition', 'attachment', filename="240608_Hamoon_34100円_出展者宛いきものづくし大規模請求書.pdf")
                msg.attach(mime_base)
            send_gmail(msg)
            # for sakka in sakkas:
                # print(f'名前: {sakka[NAME]}', f'メール: {sakka[EMAIL_COLUMN]}', f'PDF: {sakka['pdf']}')
                # print(sakka)
            print('メールを送信しました。')
        else:
            print('楽しんでいきましょう。')

root = tk.Tk()
root.title('Monetize(ξ)')
root_geo = f"{w}x{h}"
root.geometry(root_geo)
root.iconbitmap(CUR_DIR + '/icon.ico')
app = Main(root=root)
app.mainloop()